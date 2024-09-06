import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('due_Records.xlsx') #placeholder name for file 
sheet = wb.get_sheet_by_name("Sheet1")
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row= r, column=1).value
        email = sheet.cell(row = r, column=2).value 

        unpaidMembers[name] = email


smtpObj = smtplib.SMTP("smtp.example.com", 587) #placeholder smtp. Change according to email provider 
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('example_email_address@example.com', sys.argv[1])# placeholder email address 

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!'" % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('example_email_address@example.com', email, body) #placeholder email address 

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
        
smtpObj.quit()



